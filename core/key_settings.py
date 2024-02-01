from openpyxl import *
import sys
import os
from pathlib import *

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *

META_DIR=Path(os.path.abspath("./"))

KEYS={\
            "A":Qt.Key_A,\
            "B":Qt.Key_B,\
            "C":Qt.Key_C,\
            "D":Qt.Key_D,\
            "E":Qt.Key_E,\
            "F":Qt.Key_F,\
            "G":Qt.Key_G,\
            "H":Qt.Key_H,\
            "I":Qt.Key_I,\
            "J":Qt.Key_J,\
            "K":Qt.Key_K,\
            "L":Qt.Key_L,\
            "M":Qt.Key_M,\
            "N":Qt.Key_N,\
            "O":Qt.Key_O,\
            "P":Qt.Key_P,\
            "Q":Qt.Key_Q,\
            "R":Qt.Key_R,\
            "S":Qt.Key_S,\
            "T":Qt.Key_T,\
            "U":Qt.Key_U,\
            "V":Qt.Key_V,\
            "W":Qt.Key_W,\
            "X":Qt.Key_X,\
            "Y":Qt.Key_Y,\
            "Z":Qt.Key_Z\
            }

def get_current_settings():
    ws=load_workbook(filename=str(META_DIR / "key_settings.xlsx"))["key"]
    user_keys=[ws[f"B{i}"].value for i in range(2,5+1)]
    d={"Y":user_keys[0],"K":user_keys[1],"L":user_keys[2],"M":user_keys[3]}
    print(d)
    return d

class K_s_window(QWidget):
    signal=pyqtSignal()
    def __init__(self):
        super(K_s_window,self).__init__()
        self.current_keys=get_current_settings()#(Y,K,L,M)
        self.setWindowTitle("设置")

        self.lb1=QLabel("键位设置：(目前仅支持设定为26个字母键)")
        self.K=Line("相同")
        self.L=Line("不同")
        self.M=Line("特殊")
        self.Y=Line("开始")
        self.K.set_line_text(self.current_keys["K"])
        self.L.set_line_text(self.current_keys["L"])
        self.M.set_line_text(self.current_keys["M"])
        self.Y.set_line_text(self.current_keys["Y"])

        self.btn=QPushButton("确定")
        self.btn.clicked.connect(self.btn_func)   

        v_layout=QVBoxLayout()
        v_layout.addWidget(self.lb1)
        v_layout.addWidget(self.K)
        v_layout.addWidget(self.L)
        v_layout.addWidget(self.M)
        v_layout.addWidget(self.Y)
        v_layout.addWidget(self.btn)
        self.setLayout(v_layout)
    

    def write_settings(self):
        xlsxfile=str(META_DIR / "key_settings.xlsx")
        wb=load_workbook(filename=xlsxfile)
        ws=wb["key"]
        ws["B2"]=self.current_keys["Y"]
        ws["B3"]=self.current_keys["K"]
        ws["B4"]=self.current_keys["L"]
        ws["B5"]=self.current_keys["M"]

        wb.save(xlsxfile)

    def btn_func(self):
        _bak=self.current_keys.copy()
        self.current_keys["Y"]=self.Y.edit_line_text().upper()
        self.current_keys["K"]=self.K.edit_line_text().upper()
        self.current_keys["M"]=self.M.edit_line_text().upper()
        self.current_keys["L"]=self.L.edit_line_text().upper()

        if(not _bak==self.current_keys):
            self.write_settings()
            print("键位设定已经改变，可以在“key_settings.xlsx”进行设置。")

        self.signal.emit()
        self.close()

class Line(QWidget):
    def __init__(self,key):
        super(Line,self).__init__()
        self.lb=QLabel(key)
        self.edit_line=QLineEdit()
        self.edit_line.setMaxLength(1)
        h_layout=QHBoxLayout()
        h_layout.addWidget(self.lb)
        h_layout.addWidget(self.edit_line)
        self.setLayout(h_layout)

    def set_line_text(self,text):
        self.edit_line.setText(text)

    def edit_line_text(self):
        return self.edit_line.text()


if(__name__=="__main__"):
    META_DIR=Path(os.path.abspath("./")).parent
    app=QApplication(sys.argv)
    a=K_s_window()
    a.show()
    sys.exit(app.exec())
    print(KEYS["G"])
