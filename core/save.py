from openpyxl import *
import os
from pathlib import *

META_DIR=Path(os.path.abspath("./"))

class Saver():
    def __init__(self,slot="saves.xlsx"):
        self.META_DIR=META_DIR
        self.saves_dir=self.META_DIR / "saves"
        self.slot=self.saves_dir / slot

        if(not self.have_slot()):self.create_slot()

        self.wb=load_workbook(filename=self.slot)
        self.ws=self.wb["score"]

    def have_slot(self):#存档文件是否存在
        return self.slot.exists()

    def create_slot(self):#如果没有存档则从预留模版中复制创建。
        _copy=load_workbook(str(self.saves_dir / "Template.xlsx"))
        _copy.save(self.slot)

    def add_score(self,data_tuple):#(玩家ID,选用模组,正确数,总题数,用时,rating,日期,时间)
      
        #利用指针变量，找到没有数据的行
        ptr=1
        while(self.ws[f"A{ptr}"].value):ptr+=1
        print("ptr={0:n}".format(ptr))

        columns="ABCDEFGH"
        for i in range(len(columns)):
            self.ws["{0}{1}".format(columns[i],ptr)]=data_tuple[i]
        
        self.wb.save(self.slot)

if(__name__=="__main__"):
    META_DIR=META_DIR.parent
    a=Saver()
    print(a.have_slot())
    a.add_score((1,2,3,4,5,6,"2024/2/1","0:14"))
